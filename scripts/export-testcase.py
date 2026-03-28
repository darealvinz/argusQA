#!/usr/bin/env python3
"""
Argus QA — Test Case XLSX Exporter

Converts an Argus .testcase.md file into a styled Excel workbook.

Usage:
    python scripts/export-testcase.py <source.testcase.md> [output.xlsx]

If output path is omitted, writes to .argus/artifacts/<feature>.testcase.xlsx
"""

import sys
import os
import re
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ── Styling constants (matching reference Testcase_CL2025_0.1.xlsx) ──────────

FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_DATA = Font(name="Arial", size=9)
FONT_SUMMARY_LABEL = Font(name="Arial", size=11, bold=True)
FONT_SUMMARY_VALUE = Font(name="Arial", size=11)
FONT_SUMMARY_TITLE = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_SECTION_HEADER = Font(name="Arial", size=11, bold=True)

FILL_HEADER = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_TYPE_POSITIVE = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FILL_TYPE_NEGATIVE = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
FILL_PRIORITY_HIGH = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
FILL_PRIORITY_MEDIUM = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
FILL_PRIORITY_LOW = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
FILL_KNOWN_BUG = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
FILL_SUMMARY_TITLE = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
FILL_SECTION_HEADER = PatternFill(start_color="E8EAED", end_color="E8EAED", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_DATA_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_DATA_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLUMN_CONFIG = [
    # (header, width, data_alignment)
    ("ID", 14, ALIGN_DATA_LEFT),
    ("Test Case Title", 38, ALIGN_DATA_LEFT),
    ("Type", 11, ALIGN_DATA_CENTER),
    ("Test Data", 28, ALIGN_DATA_LEFT),
    ("Steps", 46, ALIGN_DATA_LEFT),
    ("Expected Result", 44, ALIGN_DATA_LEFT),
    ("Priority", 10, ALIGN_DATA_CENTER),
    ("Known Bug / Notes", 40, ALIGN_DATA_LEFT),
    ("Status", 12, ALIGN_DATA_CENTER),
]

HEADER_ROW_HEIGHT = 26
DATA_ROW_HEIGHT = 52


# ── Parsing ──────────────────────────────────────────────────────────────────

def parse_testcase_md(filepath):
    """Parse a .testcase.md file and return structured data."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    # Extract header metadata
    feature_match = re.search(r"\*\*Feature:\*\*\s*(.+)", content)
    generated_match = re.search(r"\*\*Generated from:\*\*\s*(.+)", content)
    feature_id = feature_match.group(1).strip() if feature_match else "unknown"
    generated_from = generated_match.group(1).strip() if generated_match else "unknown"

    # Parse sections and test cases
    test_cases = []
    current_layer = ""
    current_type = ""
    in_table = False

    for line in content.split("\n"):
        line = line.strip()

        # Detect layer headers
        if line.startswith("## ") and "Layer" in line:
            current_layer = line.replace("## ", "").strip()
            in_table = False
            continue

        # Detect type sub-headers
        if line.startswith("### Positive"):
            current_type = "Positive"
            in_table = False
            continue
        if line.startswith("### Negative"):
            current_type = "Negative"
            in_table = False
            continue

        # Skip table header rows and separator rows
        if line.startswith("| TestcaseID") or line.startswith("|---"):
            in_table = True
            continue

        # Skip non-table lines
        if not line.startswith("|") or not in_table:
            if line and not line.startswith("#") and not line.startswith("**") and not line.startswith("---"):
                in_table = False
            continue

        # Parse table row — handle escaped pipes in cell content
        row_content = line[1:-1] if line.endswith("|") else line[1:]
        # Replace escaped pipes with placeholder before splitting
        row_content = row_content.replace("\\|", "\x00")
        cells = [c.replace("\x00", "|").strip() for c in row_content.split("|")]

        if len(cells) < 4:
            continue

        tc_id = cells[0]
        test_step = cells[1]
        test_data = cells[2] if cells[2] else "\u2014"
        expected_result = cells[3]

        # Extract KNOWN BUG from test step or expected result
        known_bug = ""
        bug_pattern = r"\*\*KNOWN BUG:\*\*\s*(.+)"

        bug_in_step = re.search(bug_pattern, test_step)
        bug_in_result = re.search(bug_pattern, expected_result)

        if bug_in_step:
            known_bug = bug_in_step.group(1).strip()
            test_step = re.sub(r"\s*\*\*KNOWN BUG:\*\*.*", "", test_step).strip()
        if bug_in_result:
            known_bug = bug_in_result.group(1).strip()
            expected_result = re.sub(r"\s*\*\*KNOWN BUG:\*\*.*", "", expected_result).strip()

        # Derive short title from test step
        title = derive_title(test_step)

        # Determine priority
        priority = determine_priority(tc_id, current_layer, current_type, test_step)

        test_cases.append({
            "id": tc_id,
            "title": title,
            "type": current_type,
            "test_data": test_data,
            "steps": test_step,
            "expected_result": expected_result,
            "priority": priority,
            "known_bug": known_bug,
            "status": "",
            "layer": current_layer,
        })

    return feature_id, generated_from, test_cases


def derive_title(test_step):
    """Create a short title from the full test step text."""
    # Remove common prefixes
    title = test_step
    for prefix in ["Verify ", "Navigate to ", "Click ", "Submit ", "Attempt ", "Type "]:
        if title.startswith(prefix):
            title = title[len(prefix):]
            break

    # Truncate at first major punctuation or conjunction
    for sep in [" and verify", " and check", " and observe", ". "]:
        idx = title.lower().find(sep)
        if idx > 0:
            title = title[:idx]
            break

    # Cap length
    if len(title) > 60:
        title = title[:57] + "..."

    return title


def determine_priority(tc_id, layer, tc_type, test_step):
    """Determine priority based on layer, content, and test ID."""
    step_lower = test_step.lower()

    # Security layer is always High
    if "Security" in layer:
        return "High"

    # Known bugs are High
    if "KNOWN BUG" in test_step:
        return "High"

    # Login, purchase, form submission, cookie consent = High
    high_keywords = ["login", "log in", "submit", "buy now", "purchase", "cookie consent",
                     "password", "credential", "xss", "sql injection", "brute force",
                     "javascript error", "search", "currency"]
    if any(kw in step_lower for kw in high_keywords):
        return "High"

    # Core page elements (first ~7 test cases in UI) = High
    try:
        num = int(re.search(r"(\d+)$", tc_id).group(1))
        if num <= 7 and "UI" in layer:
            return "High"
    except (AttributeError, ValueError):
        pass

    # Footer, sponsor, delivery logos = Low
    low_keywords = ["footer", "sponsor", "delivery logo", "payment logo",
                    "company location", "social media icon", "partner logo"]
    if any(kw in step_lower for kw in low_keywords):
        return "Low"

    # Default to Medium
    return "Medium"


# ── Excel generation ─────────────────────────────────────────────────────────

def create_summary_sheet(ws, feature_id, generated_from, test_cases):
    """Populate the Summary sheet."""
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    total = len(test_cases)
    high = sum(1 for tc in test_cases if tc["priority"] == "High")
    medium = sum(1 for tc in test_cases if tc["priority"] == "Medium")
    low = sum(1 for tc in test_cases if tc["priority"] == "Low")
    positive = sum(1 for tc in test_cases if tc["type"] == "Positive")
    negative = sum(1 for tc in test_cases if tc["type"] == "Negative")
    passed = sum(1 for tc in test_cases if tc["status"].lower() == "pass")
    failed = sum(1 for tc in test_cases if tc["status"].lower() == "fail")
    not_run = total - passed - failed

    rows = [
        ("Feature", feature_id, "title"),
        ("Export Date", date.today().isoformat(), None),
        ("Generated From", generated_from, None),
        (None, None, None),
        ("Total Test Cases", total, None),
        (None, None, None),
        ("Priority", "Count", "section"),
        ("High", high, None),
        ("Medium", medium, None),
        ("Low", low, None),
        (None, None, None),
        ("Type", "Count", "section"),
        ("Positive", positive, None),
        ("Negative", negative, None),
        (None, None, None),
        ("Status", "Count", "section"),
        ("Pass", passed, None),
        ("Fail", failed, None),
        ("Not Run", not_run, None),
    ]

    for row_idx, (label, value, style) in enumerate(rows, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        cell_b = ws.cell(row=row_idx, column=2, value=value)

        if style == "title":
            cell_a.font = FONT_SUMMARY_TITLE
            cell_b.font = FONT_SUMMARY_TITLE
            cell_a.fill = FILL_SUMMARY_TITLE
            cell_b.fill = FILL_SUMMARY_TITLE
            cell_a.alignment = Alignment(vertical="center")
            cell_b.alignment = Alignment(vertical="center")
        elif style == "section":
            cell_a.font = FONT_SECTION_HEADER
            cell_b.font = FONT_SECTION_HEADER
            cell_a.fill = FILL_SECTION_HEADER
            cell_b.fill = FILL_SECTION_HEADER
            cell_a.alignment = Alignment(vertical="center")
            cell_b.alignment = Alignment(vertical="center")
        elif label is not None:
            cell_a.font = FONT_SUMMARY_LABEL
            cell_b.font = FONT_SUMMARY_VALUE
            cell_a.alignment = Alignment(vertical="center")
            cell_b.alignment = Alignment(vertical="center")

        if label is not None:
            cell_a.border = THIN_BORDER
            cell_b.border = THIN_BORDER


def create_testcases_sheet(ws, test_cases):
    """Populate the Test Cases sheet with styled data."""

    # Set column widths and write headers
    for col_idx, (header, width, _) in enumerate(COLUMN_CONFIG, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = THIN_BORDER

    # Header row height
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:I{len(test_cases) + 1}"

    # Write data rows
    for row_offset, tc in enumerate(test_cases):
        row_idx = row_offset + 2
        is_even = row_offset % 2 == 0
        base_fill = FILL_ROW_EVEN if is_even else FILL_ROW_ODD

        values = [
            tc["id"],
            tc["title"],
            tc["type"],
            tc["test_data"],
            tc["steps"],
            tc["expected_result"],
            tc["priority"],
            tc["known_bug"],
            tc["status"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_DATA
            cell.border = THIN_BORDER
            cell.alignment = COLUMN_CONFIG[col_idx - 1][2]

            # Apply base alternating row fill
            cell.fill = base_fill

        # Override Type column fill (col 3)
        type_cell = ws.cell(row=row_idx, column=3)
        if tc["type"] == "Positive":
            type_cell.fill = FILL_TYPE_POSITIVE
        elif tc["type"] == "Negative":
            type_cell.fill = FILL_TYPE_NEGATIVE

        # Override Priority column fill (col 7)
        priority_cell = ws.cell(row=row_idx, column=7)
        if tc["priority"] == "High":
            priority_cell.fill = FILL_PRIORITY_HIGH
        elif tc["priority"] == "Medium":
            priority_cell.fill = FILL_PRIORITY_MEDIUM
        elif tc["priority"] == "Low":
            priority_cell.fill = FILL_PRIORITY_LOW

        # Override Known Bug column fill (col 8)
        if tc["known_bug"]:
            ws.cell(row=row_idx, column=8).fill = FILL_KNOWN_BUG

        # Set row height
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/export-testcase.py <source.testcase.md> [output.xlsx]")
        print()
        print("Examples:")
        print("  python scripts/export-testcase.py .argus/test-cases/login.testcase.md")
        print("  python scripts/export-testcase.py .argus/test-cases/login.testcase.md output.xlsx")
        sys.exit(1)

    source_path = sys.argv[1]

    if not os.path.exists(source_path):
        print(f"Error: Source file not found: {source_path}")
        sys.exit(1)

    # Determine output path
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        feature_id = os.path.basename(source_path).replace(".testcase.md", "")
        artifacts_dir = os.path.join(os.path.dirname(os.path.dirname(source_path)), "artifacts")
        if not os.path.exists(os.path.join(os.path.dirname(os.path.dirname(source_path)), "artifacts")):
            # Fall back to .argus/artifacts relative to CWD
            artifacts_dir = os.path.join(".argus", "artifacts")
        os.makedirs(artifacts_dir, exist_ok=True)
        output_path = os.path.join(artifacts_dir, f"{feature_id}.testcase.xlsx")

    # Parse source
    print(f"Reading: {source_path}")
    feature_id, generated_from, test_cases = parse_testcase_md(source_path)
    print(f"Found {len(test_cases)} test cases for feature: {feature_id}")

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    create_summary_sheet(ws_summary, feature_id, generated_from, test_cases)

    # Sheet 2: Test Cases
    ws_cases = wb.create_sheet("Test Cases")
    create_testcases_sheet(ws_cases, test_cases)

    # Save
    wb.save(output_path)
    print(f"Exported: {output_path}")
    print()

    # Print summary
    high = sum(1 for tc in test_cases if tc["priority"] == "High")
    medium = sum(1 for tc in test_cases if tc["priority"] == "Medium")
    low = sum(1 for tc in test_cases if tc["priority"] == "Low")
    positive = sum(1 for tc in test_cases if tc["type"] == "Positive")
    negative = sum(1 for tc in test_cases if tc["type"] == "Negative")
    bugs = sum(1 for tc in test_cases if tc["known_bug"])

    print(f"  Total:    {len(test_cases)}")
    print(f"  Priority: {high} High / {medium} Medium / {low} Low")
    print(f"  Type:     {positive} Positive / {negative} Negative")
    print(f"  Bugs:     {bugs} known bugs flagged")


if __name__ == "__main__":
    main()
